from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl

# Load test data from Excel
def load_test_data(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    data = []
    for row in range(2, sheet.max_row + 1):
        data.append({
            "name": sheet.cell(row, 1).value,
            "email": sheet.cell(row, 2).value,
            "password": sheet.cell(row, 3).value,
            "confirm_password": sheet.cell(row, 4).value,
        })
    return data

# Initialize WebDriver
driver = webdriver.Chrome()

try:
    # Load test data
    test_data = load_test_data("test_data.xlsx")
    
    for data in test_data:
        driver.get("https://example.com/register")  # Replace with actual URL
        
        # Fill out the form
        driver.find_element(By.ID, "name").send_keys(data["name"])
        driver.find_element(By.ID, "email").send_keys(data["email"])
        driver.find_element(By.ID, "password").send_keys(data["password"])
        driver.find_element(By.ID, "confirm_password").send_keys(data["confirm_password"])
        
        # Submit the form
        driver.find_element(By.ID, "submit").click()
        
        # Capture results
        result = driver.find_element(By.ID, "result").text
        print(f"Test case for {data['email']} resulted in: {result}")
    
finally:
    driver.quit()
